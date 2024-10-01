from openpyxl import load_workbook
from utils import add_space_after_nechet_ned, find_closest_sheet_name
import openpyxl

def wb_name():
    wb = load_workbook("1S3kj0zo_QDERJu7O2QU1J4gMRx-K381m.xlsx")
    return [add_space_after_nechet_ned(item.strip()) for item in wb.sheetnames]

def get_schedule(day_name, schedule_name):
    wb = load_workbook("1S3kj0zo_QDERJu7O2QU1J4gMRx-K381m.xlsx")
    closest_sheet_name = find_closest_sheet_name(wb.sheetnames, schedule_name)
    if not closest_sheet_name:
        return "Расписание не найдено"
    ws = wb[closest_sheet_name]

    def column_letter_to_number(column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)

    # Определение диапазона строк и столбцов
    min_row = 8
    max_row = 43
    min_col_FV = column_letter_to_number('FV')
    min_col_FX = column_letter_to_number('FX')

    # Чтение данных из столбцов FV и FX и создание списка списков
    data_combined = []

    for row in range(min_row, max_row + 1):
        fv_cell = ws.cell(row=row, column=min_col_FV)
        fx_cell = ws.cell(row=row, column=min_col_FX)

        fv_value = fv_cell.value if fv_cell.value is not None else "Не указано"
        fx_value = fx_cell.value if fx_cell.value is not None else "Не указано"

        data_combined.append([fv_value, fx_value])

    # Функция для разбиения данных на блоки по 6 записей
    def split_into_days(data, num_pairs_per_day):
        return [data[i:i + num_pairs_per_day] for i in range(0, len(data), num_pairs_per_day)]

    # Разделение данных на блоки по 6 пар
    data_by_days = split_into_days(data_combined, 6)

    # Функция для форматирования данных в нужный формат
    def format_schedule(day_data, day_name):
        formatted_schedule = []
        times = ["08:00 — 09:35", "09:45 — 11:20", "11:50 — 13:25", "13:55 — 15:30", "15:40 — 17:15", "17:25 — 19:00"]
        emojis = ["1️⃣", "2️⃣", "3️⃣", "4️⃣", "5️⃣", "6️⃣"]

        for index, (subject, teacher) in enumerate(day_data):
            time = times[index] if index < len(times) else "Не указано"
            emoji = emojis[index] if index < len(emojis) else ""
            formatted_schedule.append(f"{emoji} {time} : {subject} ({teacher})")

        return f"Расписание на {day_name}:\n" + "\n".join(formatted_schedule)

    # Дни недели для расписания
    days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]

    # Проверка, что day_name входит в список дней недели
    if day_name not in days_of_week:
        return "Неправильный день недели"

    day_index = days_of_week.index(day_name)
    day_data = data_by_days[day_index] if day_index < len(data_by_days) else []

    return format_schedule(day_data, day_name)